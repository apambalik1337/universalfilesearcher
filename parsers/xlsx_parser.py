"""
xlsx_parser.py

Parses modern Excel files (.xlsx, .xlsm) using openpyxl.

Requirements from spec (Priority 1 - MUST WORK):
  - read every worksheet (not just the first/active one)
  - read every populated cell
  - preserve worksheet name and cell coordinate
  - handle empty cells safely
  - handle formulas appropriately (index the computed value, not the formula
    text, so a search for a name finds it even if the cell contains a
    formula that produces that name)
  - handle merged cells without crashing
"""

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .base_parser import BaseParser, ParseResult, Record


class XlsxParser(BaseParser):
    extensions = (".xlsx", ".xlsm")
    display_name = "Excel (.xlsx)"

    def parse(self, file_path: str) -> ParseResult:
        records = []
        try:
            # data_only=True -> read the last computed value of formulas
            # instead of the formula string itself, per spec requirement.
            wb = load_workbook(file_path, data_only=True, read_only=True)
        except InvalidFileException as e:
            return ParseResult(success=False, error=f"Not a valid xlsx file: {e}")
        except Exception as e:
            return ParseResult(success=False, error=f"Failed to open workbook: {e}")

        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                try:
                    for row in ws.iter_rows():
                        for cell in row:
                            value = cell.value
                            if value is None:
                                continue  # empty cell, skip safely
                            text = str(value).strip()
                            if not text:
                                continue
                            records.append(
                                Record(
                                    content=text,
                                    sheet=sheet_name,
                                    cell=cell.coordinate,
                                )
                            )
                except Exception as e:
                    # A problem in one sheet shouldn't lose the others.
                    records.append(
                        Record(
                            content=f"[Warning: could not fully read sheet '{sheet_name}': {e}]",
                            sheet=sheet_name,
                        )
                    )
        finally:
            wb.close()

        if not records:
            return ParseResult(success=True, records=[], warning="No content found in workbook")

        return ParseResult(success=True, records=records)

    def is_dependency_available(self) -> bool:
        try:
            import openpyxl  # noqa
            return True
        except ImportError:
            return False

    def dependency_message(self) -> str:
        return "Install with: pip install openpyxl"
